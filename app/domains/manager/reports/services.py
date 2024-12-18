from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Side
from datetime import date
from io import BytesIO
from urllib.parse import quote

# DAOs
from app.domains.tasks.dao import TaskDAO, TaskAssignmentDAO
from app.domains.projects.dao import ProjectDAO, ProjectMemberDAO
from app.domains.users.dao import UserDAO

# Схемы
from app.domains.users.schemas import UserDB

THIN_BORDER = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin")
)

BOLD_BORDER = Border(
    left=Side(border_style="thick"),
    right=Side(border_style="thick"),
    top=Side(border_style="thick"),
    bottom=Side(border_style="thick")
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


class ReportsService:
    @staticmethod
    async def create_report_by_task(
            task_id: int,
            current_user: UserDB
    ):
        if current_user.role.name != "Менеджер":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Пользователь не является менеджером"
            )
        task = await TaskDAO.get_task_with_project(task_id)
        if not task:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Задача не найдена"
            )
        users = await TaskAssignmentDAO.get_task_assignments(task_id)
        wb = Workbook()
        ws = wb.active
        ws.title = f"Отчёт по задаче"

        # Информация о задаче
        ws.merge_cells('A1:B1')
        ws['A1'].border = BOLD_BORDER
        ws['A1'].alignment = ALIGN_CENTER
        ws['B1'].border = BOLD_BORDER
        ws['A1'] = "Задача"
        task_data = [
            ("Название задачи", task.title),
            ("Описание", task.description or "Не указано"),
            ("Дата начала", task.start_date.strftime("%d.%m.%Y")),
            ("Дата завершения", task.due_date.strftime("%d.%m.%Y")),
            ("Статус", task.status.name),
            ("Приоритет", task.priority.name),
        ]
        for i, (header, value) in enumerate(task_data):
            ws.cell(row=i + 2, column=1).value = header
            ws.cell(row=i + 2, column=1).border = THIN_BORDER
            ws.cell(row=i + 2, column=2).value = value
            ws.cell(row=i + 2, column=2).border = THIN_BORDER

        # Информация о проекте
        ws.merge_cells('D1:E1')
        ws['D1'] = "Проект, в которой задача"
        ws['D1'].border = BOLD_BORDER
        ws['E1'].border = BOLD_BORDER
        ws['D1'].alignment = ALIGN_CENTER
        if task.project:
            project_data = [
                ("Название проекта", task.project.title),
                ("Описание", task.project.description or "Не указано"),
                ("Дата начала", task.project.start_date.strftime("%d.%m.%Y")),
                ("Дата завершения", task.project.due_date.strftime("%d.%m.%Y")),
                ("Статус", task.project.status.name),
            ]
            for i, (header, value) in enumerate(project_data):
                ws.cell(row=i + 2, column=4).value = header
                ws.cell(row=i + 2, column=4).border = THIN_BORDER
                ws.cell(row=i + 2, column=5).value = value
                ws.cell(row=i + 2, column=5).border = THIN_BORDER
        else:
            ws.merge_cells('D2:E2')
            ws['D2'] = "Задача не привязана к проекту"
            ws['D2'].alignment = ALIGN_CENTER
            ws['D2'].border = THIN_BORDER
            ws['E2'].border = THIN_BORDER

        # Информация о пользователях
        ws.merge_cells('A9:D9')
        ws['A9'] = "Участники задачи"
        ws['A9'].border = BOLD_BORDER
        ws['B9'].border = BOLD_BORDER
        ws['C9'].border = BOLD_BORDER
        ws['D9'].border = BOLD_BORDER
        ws['A9'].alignment = ALIGN_CENTER
        if users:
            ws['A10'] = "Имя"
            ws['A10'].border = BOLD_BORDER
            ws['B10'] = "Фамилия"
            ws['B10'].border = BOLD_BORDER
            ws['C10'] = "Отчество"
            ws['C10'].border = BOLD_BORDER
            ws['D10'] = "Должность"
            ws['D10'].border = BOLD_BORDER
            row = 11
            for user in users:
                user_data = [
                    user.user.first_name,
                    user.user.last_name,
                    user.user.patronymic if user.user.patronymic else "-",
                    user.user.position.name if user.user.position else "Не указана",
                ]
                for i, value in enumerate(user_data):
                    ws.cell(row=row, column=i + 1).value = value
                    ws.cell(row=row, column=i + 1).border = THIN_BORDER
                row += 1
        else:
            ws.merge_cells('A9:D9')
            ws['A9'] = "Нет участников задачи"
            ws['A9'].border = THIN_BORDER
            ws['B9'].border = THIN_BORDER
            ws['C9'].border = THIN_BORDER
            ws['D9'].border = THIN_BORDER
            ws['A9'].alignment = ALIGN_CENTER

        # Сохранение отчёта
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Имя файла
        filename = f"Report_task_{task.title}_{date.today().isoformat()}"
        encoded_file_name = quote(filename)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_file_name}"}
        )

    @staticmethod
    async def create_report_by_user(
            user_id: int,
            current_user: UserDB
    ):
        if current_user.role.name != "Менеджер":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Пользователь не является менеджером"
            )
        user = await UserDAO.find_by_id(user_id)
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Пользователь не найден"
            )
        tasks = await TaskAssignmentDAO.get_user_tasks(user_id)

        projects = await ProjectMemberDAO.get_user_projects(user_id)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Отчёт по пользователю"

        # Информация о пользователе
        ws.merge_cells('A1:B1')
        ws['A1'] = "Пользователь"
        ws['A1'].border = BOLD_BORDER
        ws['B1'].border = BOLD_BORDER
        ws['A1'].alignment = ALIGN_CENTER
        user_data = [
            ("Имя", user.first_name),
            ("Фамилия", user.last_name),
            ("Отчество", user.patronymic if user.patronymic else "-"),
            ("Должность", user.position.name if user.position else "Не указана"),
        ]
        for i, (header, value) in enumerate(user_data):
            ws.cell(row=i + 2, column=1).value = header
            ws.cell(row=i + 2, column=1).border = THIN_BORDER
            ws.cell(row=i + 2, column=2).value = value
            ws.cell(row=i + 2, column=2).border = THIN_BORDER

        # Задачи пользователя
        ws.merge_cells('A8:D8')
        ws['A8'] = "Задачи пользователя"
        ws['A8'].border = BOLD_BORDER
        ws['B8'].border = BOLD_BORDER
        ws['C8'].border = BOLD_BORDER
        ws['D8'].border = BOLD_BORDER
        ws['A8'].alignment = ALIGN_CENTER
        if tasks:
            ws['A9'] = "Название задачи"
            ws['A9'].border = BOLD_BORDER
            ws['B9'] = "Статус"
            ws['B9'].border = BOLD_BORDER
            ws['C9'] = "Дата начала"
            ws['C9'].border = BOLD_BORDER
            ws['D9'] = "Дата завершения"
            ws['D9'].border = BOLD_BORDER
            row = 10
            for task in tasks:
                task_data = [
                    task.task.title,
                    task.task.status.name,
                    task.task.start_date.strftime("%d.%m.%Y"),
                    task.task.due_date.strftime("%d.%m.%Y"),
                ]
                for i, value in enumerate(task_data):
                    ws.cell(row=row, column=i + 1).value = value
                    ws.cell(row=row, column=i + 1).border = THIN_BORDER
                row += 1
        else:
            ws.merge_cells('A10:D10')
            ws['A10'] = "Нет задач у пользователя"
            ws['A10'].border = THIN_BORDER
            ws['B10'].border = THIN_BORDER
            ws['C10'].border = THIN_BORDER
            ws['D10'].border = THIN_BORDER
            ws['A10'].alignment = ALIGN_CENTER

        # Проекты пользователя
        ws.merge_cells('F8:I8')
        ws['F8'] = "Проекты пользователя"
        ws['F8'].border = BOLD_BORDER
        ws['G8'].border = BOLD_BORDER
        ws['H8'].border = BOLD_BORDER
        ws['I8'].border = BOLD_BORDER
        ws['F8'].alignment = ALIGN_CENTER
        if projects:
            ws['F9'] = "Название проекта"
            ws['F9'].border = BOLD_BORDER
            ws['G9'] = "Описание"
            ws['G9'].border = BOLD_BORDER
            ws['H9'] = "Дата начала"
            ws['H9'].border = BOLD_BORDER
            ws['I9'] = "Дата завершения"
            ws['I9'].border = BOLD_BORDER
            row = 10
            for project in projects:
                project_data = [
                    project.project.title,
                    project.project.description or "Не указано",
                    project.project.start_date.strftime("%d.%m.%Y"),
                    project.project.due_date.strftime("%d.%m.%Y"),
                ]
                for i, value in enumerate(project_data):
                    ws.cell(row=row, column=i + 6).value = value
                    ws.cell(row=row, column=i + 6).border = THIN_BORDER
                row += 1
        else:
            ws.merge_cells('F9:I9')
            ws['F9'] = "Нет проектов у пользователя"
            ws['F9'].border = THIN_BORDER
            ws['G9'].border = THIN_BORDER
            ws['H9'].border = THIN_BORDER
            ws['I9'].border = THIN_BORDER
            ws['F9'].alignment = ALIGN_CENTER

        # Сохранение отчёта
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Имя файла
        filename = f"Report_user_{user.first_name}_{user.last_name}_{date.today().isoformat()}"
        encoded_file_name = quote(filename)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_file_name}"}
        )

    @staticmethod
    async def create_report_by_project(
            project_id: int,
            current_user: UserDB
    ):
        if current_user.role.name != "Менеджер":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Пользователь не является менеджером"
            )
        project = await ProjectDAO.find_by_id(project_id)
        if not project:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Проект не найден"
            )
        users = await ProjectMemberDAO.get_project_members(project_id)
        tasks = await ProjectDAO.get_project_tasks(project_id)
        tasks = tasks.tasks

        wb = Workbook()
        ws = wb.active
        ws.title = f"Отчёт по проекту"

        # Информация о проекте
        ws.merge_cells('A1:B1')
        ws['A1'] = "Проект"
        ws['A1'].border = BOLD_BORDER
        ws['B1'].border = BOLD_BORDER
        ws['A1'].alignment = ALIGN_CENTER
        project_data = [
            ("Название проекта", project.title),
            ("Описание", project.description or "Не указано"),
            ("Дата начала", project.start_date.strftime("%d.%m.%Y")),
            ("Дата завершения", project.due_date.strftime("%d.%m.%Y")),
            ("Статус", project.status.name),
        ]
        for i, (header, value) in enumerate(project_data):
            ws.cell(row=i + 2, column=1).value = header
            ws.cell(row=i + 2, column=1).border = THIN_BORDER
            ws.cell(row=i + 2, column=2).value = value
            ws.cell(row=i + 2, column=2).border = THIN_BORDER

        # Участники проекта
        ws.merge_cells('A8:D8')
        ws['A8'] = "Участники задачи"
        ws['A8'].border = BOLD_BORDER
        ws['B8'].border = BOLD_BORDER
        ws['C8'].border = BOLD_BORDER
        ws['D8'].border = BOLD_BORDER
        ws['A8'].alignment = ALIGN_CENTER
        if users:
            ws['A9'] = "Имя"
            ws['A9'].border = BOLD_BORDER
            ws['B9'] = "Фамилия"
            ws['B9'].border = BOLD_BORDER
            ws['C9'] = "Отчество"
            ws['C9'].border = BOLD_BORDER
            ws['D9'] = "Должность"
            ws['D9'].border = BOLD_BORDER
            row = 10
            for user in users:
                user_data = [
                    user.user.first_name,
                    user.user.last_name,
                    user.user.patronymic if user.user.patronymic else "-",
                    user.user.position.name if user.user.position else "Не указана",
                ]
                for i, value in enumerate(user_data):
                    ws.cell(row=row, column=i + 1).value = value
                    ws.cell(row=row, column=i + 1).border = THIN_BORDER
                row += 1
        else:
            ws.merge_cells('A9:D9')
            ws['A9'] = "Нет участников проекта"
            ws['A9'].border = THIN_BORDER
            ws['B9'].border = THIN_BORDER
            ws['C9'].border = THIN_BORDER
            ws['D9'].border = THIN_BORDER
            ws['A9'].alignment = ALIGN_CENTER

        # Задачи проекта
        ws.merge_cells('F8:K8')
        ws['F8'] = "Задачи проекта"
        ws['F8'].border = BOLD_BORDER
        ws['G8'].border = BOLD_BORDER
        ws['H8'].border = BOLD_BORDER
        ws['I8'].border = BOLD_BORDER
        ws['J8'].border = BOLD_BORDER
        ws['K8'].border = BOLD_BORDER
        ws['F8'].alignment = ALIGN_CENTER
        if tasks:
            ws['F9'] = "Название"
            ws['F9'].border = BOLD_BORDER
            ws['G9'] = "Описание"
            ws['G9'].border = BOLD_BORDER
            ws['H9'] = "Статус"
            ws['H9'].border = BOLD_BORDER
            ws['I9'] = "Дата начала"
            ws['I9'].border = BOLD_BORDER
            ws['J9'] = "Дата завершения"
            ws['J9'].border = BOLD_BORDER
            ws['K9'] = "Приоритет"
            ws['K9'].border = BOLD_BORDER
            row = 10
            for task in tasks:
                task_data = [
                    task.title,
                    task.description or "Не указано",
                    task.status.name,
                    task.start_date.strftime("%d.%m.%Y"),
                    task.due_date.strftime("%d.%m.%Y"),
                    task.priority.name,
                ]
                for i, value in enumerate(task_data):
                    ws.cell(row=row, column=i + 6).value = value
                    ws.cell(row=row, column=i + 6).border = THIN_BORDER
                row += 1
        else:
            ws.merge_cells('F9:K9')
            ws['F9'] = "Нет задач внутри проекта"
            ws['F9'].border = THIN_BORDER
            ws['G9'].border = THIN_BORDER
            ws['H9'].border = THIN_BORDER
            ws['I9'].border = THIN_BORDER
            ws['J9'].border = THIN_BORDER
            ws['K9'].border = THIN_BORDER
            ws['F9'].alignment = ALIGN_CENTER

        # Сохранение отчёта
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Имя файла
        filename = f"Report_project_{project.title}_{date.today().isoformat()}"
        encoded_file_name = quote(filename)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_file_name}"}
        )
